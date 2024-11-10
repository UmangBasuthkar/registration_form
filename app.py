from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)

# Ensure the 'logs' directory exists to save Excel files
if not os.path.exists('logs'):
    os.makedirs('logs')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/log', methods=['POST'])
def log_entry():
    name = request.form['name']
    phone = request.form['phone']
    action = request.form['action']
    date = datetime.now().strftime('%Y-%m-%d')
    timestamp = datetime.now().strftime('%I:%M:%S %p')  # 12-hour format with AM/PM
    filename = f'logs/{date}_office_log.xlsx'

    # Load or create the Excel workbook
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Set header row if new workbook
        sheet.append(["Name", "Phone", "Entry", "Exit"])

    # Process entry and exit actions
    if action == 'entry':
        print(f"Logging entry for {name} at {timestamp}")
        sheet.append([name, phone, timestamp, ''])
    elif action == 'exit':
        # Find the matching entry to update exit time
        updated = False  # Flag to track if an update occurs
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == name and row[1].value == phone and not row[3].value:
                row[3].value = timestamp  # Update the exit time in the correct cell
                updated = True
                print(f"Exit time updated for {name} at {timestamp}")
                break
        if not updated:
            print(f"No matching entry found for {name} with phone {phone} to update exit time.")

    # Save the workbook after modifications
    workbook.save(filename)
    print(f"Workbook saved: {filename}")
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
